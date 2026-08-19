"""Turn source.xlsx (a download of the Google Sheet) into dataset.json.

    python parse_sheet.py

Only needs rerunning when the sheet itself changes. A merged cell in Excel
carries its value on the first row only, so a row with an empty first cell is
read as another sense of the entry above it.
"""
import json
import os
import re

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SRC = os.path.join(HERE, 'source.xlsx')

IPA_RE = re.compile(r'/[^/\n]{1,80}/')
POS_RE = re.compile(r'^[a-zA-Z][a-zA-Z,. /]{0,24}$')
EG_SPLIT = re.compile(r'\n\s*[-–—]\s*')


def txt(v):
    """Cell value as clean text, with line endings normalised."""
    return '' if v is None else str(v).replace('\r\n', '\n').replace('\r', '\n').strip()


def flat(s):
    return re.sub(r'[ \t]+', ' ', txt(s).replace('\n', ' ')).strip()


def sense(def_cell, vi_cell):
    """Split examples (lines starting with a dash) off from the definition."""
    parts = EG_SPLIT.split(txt(def_cell))
    return {
        'def': flat(parts[0]),
        'eg': [flat(p) for p in parts[1:] if flat(p)],
        'vi': flat(vi_cell),
    }


def parse_head(cell):
    """'aardvark (n)\\n/ˈɑːd.vɑːk/\\n= initio' -> (word, pos, ipa, note)"""
    cell = txt(cell)
    ipa = ''
    m = IPA_RE.search(cell)
    if m:
        ipa = m.group(0).strip()
        cell = cell[:m.start()] + '\n' + cell[m.end():]
    lines = [l.strip() for l in cell.split('\n') if l.strip()]
    first = lines[0] if lines else ''
    pos = ''
    m = re.search(r'\(([^()]{1,25})\)\s*$', first)
    if m and POS_RE.match(m.group(1).strip()):
        pos = m.group(1).strip()
        first = first[:m.start()].strip()
    note = ' '.join(lines[1:]).strip(' -=')
    return flat(first), pos, ipa, flat(note)


entries = []
readings = []


def add(**kw):
    kw.setdefault('note', '')
    kw.setdefault('pos', '')
    kw.setdefault('ipa', '')
    kw['senses'] = [s for s in kw['senses'] if s['def'] or s['vi']]
    if kw['word'] and kw['senses']:
        entries.append(kw)


def blocks(ws, key_cols, skip_header=1):
    """Group consecutive rows into one entry: a row whose key cell is empty is
    another sense of the entry above it (the trace of a merged cell)."""
    cur, buf = None, []
    for n, row in enumerate(ws.iter_rows(values_only=True)):
        if n < skip_header:
            continue
        row = list(row) + [None] * 8
        key = tuple(txt(row[c]) for c in key_cols)
        if not any(txt(c) for c in row):
            continue
        if any(key):
            if buf:
                yield cur, buf
            cur, buf = key, [row]
        elif buf:
            buf.append(row)
    if buf:
        yield cur, buf


wb = openpyxl.load_workbook(SRC, read_only=True, data_only=True)

# --- Vocabulary: A = word (+ pos + phonetics), B = definition, C = meaning ---
for key, rows in blocks(wb['Vocabulary'], [0], skip_header=1):
    head = rows[0]
    # letter divider row: a single letter in column A and no definition
    if not txt(head[1]) and not txt(head[2]) and len(txt(head[0])) <= 2:
        continue
    word, pos, ipa, note = parse_head(head[0])
    add(type='word', word=word, pos=pos, ipa=ipa, note=note,
        senses=[sense(r[1], r[2]) for r in rows])

# --- Phrasal Verb: A = verb, B = particle, C = explain, D = meaning ---
# The verb cell is merged across all of its particles, so the current verb has
# to be carried forward: a new particle starts a new entry, both cells empty
# means another sense of the entry in progress.
def flush_phrasal(verb, part, rows):
    if rows and verb:
        add(type='phrasal', word=flat(verb + ' ' + part), verb=flat(verb), particle=flat(part),
            senses=[sense(r[2], r[3]) for r in rows])


cur_verb, cur_part, buf = '', '', []
for n, row in enumerate(wb['Phrasal Verb'].iter_rows(values_only=True)):
    if n < 1:
        continue
    row = list(row) + [None] * 8
    a, b = txt(row[0]), txt(row[1])
    if not any(txt(c) for c in row[:4]):
        continue
    if a or b:
        flush_phrasal(cur_verb, cur_part, buf)
        if a:
            cur_verb = a
        cur_part = b
        buf = [row]
    else:
        buf.append(row)
flush_phrasal(cur_verb, cur_part, buf)

# --- Idioms / Common: A = headword, B = explain, C = meaning ---
for sheet, kind in (('Idioms', 'idiom'), ('Common', 'expression')):
    for key, rows in blocks(wb[sheet], [0], skip_header=1):
        word, pos, ipa, note = parse_head(rows[0][0])
        add(type=kind, word=word, pos=pos, ipa=ipa, note=note,
            senses=[sense(r[1], r[2]) for r in rows])

# --- Grammar: A = group number, B = word, C = explain, D = meaning ---
group = ''
for n, row in enumerate(wb['Grammar'].iter_rows(values_only=True)):
    if n < 1:
        continue
    row = list(row) + [None] * 8
    if txt(row[0]):
        group = txt(row[0]).replace('.0', '')
    if txt(row[1]):
        add(type='compare', word=flat(txt(row[1])), group=group,
            senses=[sense(row[2], row[3])])

def label_paragraphs(paras):
    """Some passages are the IELTS sort, with paragraphs lettered A, B, C…
    The letter is glued to the text and sometimes doubled ("A A In 1977"). It
    only counts as a label when the letters actually run in order down the
    passage — otherwise a paragraph opening with the article "A" would lose it.
    """
    letters = [chr(ord('A') + i) for i in range(len(paras))]
    runs = sum(1 for i, p in enumerate(paras)
               if i < 26 and re.match(letters[i] + r'(\s|$)', p))
    if len(paras) < 3 or runs < max(3, int(len(paras) * 0.6)):
        return [{'text': p} for p in paras]

    out = []
    for i, p in enumerate(paras):
        L = letters[i] if i < 26 else None
        m = re.match(L + r'\s+(?:' + L + r'\s+)?(.*)$', p, re.S) if L else None
        if m and m.group(1).strip():
            out.append({'mark': L, 'text': m.group(1).strip()})
        else:
            out.append({'text': p})
    return out


# --- Reading Passage: two rows per piece, the title then the body ---
pend = None
for n, row in enumerate(wb['Reading Passage'].iter_rows(values_only=True)):
    if n < 1:
        continue
    row = list(row) + [None] * 4
    body = txt(row[1])
    if not body:
        continue
    if txt(row[0]):
        pend = {'index': txt(row[0]).replace('.0', ''), 'title': flat(body)}
    elif pend:
        pend['paras'] = label_paragraphs([flat(p) for p in body.split('\n') if flat(p)])
        readings.append(pend)
        pend = None

for n, e in enumerate(entries):
    e['id'] = 's' + str(n)

data = {'entries': entries, 'readings': readings}
json.dump(data, open(os.path.join(HERE, 'dataset.json'), 'w', encoding='utf-8'),
          ensure_ascii=False, separators=(',', ':'))

# A raw snapshot of each tab, so the tests can compare what Apps Script
# (sheet-sync.gs) reads against what this script reads. Derived, not committed.
grids = {}
for name in ('Vocabulary', 'Phrasal Verb', 'Idioms', 'Common', 'Grammar'):
    ws = wb[name]
    g = []
    for row in ws.iter_rows(values_only=True):
        g.append([('' if c is None else str(c)) for c in (list(row) + [None] * 4)[:4]])
    grids[name] = g
test_dir = os.path.join(HERE, 'test')
if os.path.isdir(test_dir):
    json.dump(grids, open(os.path.join(test_dir, 'grids.json'), 'w', encoding='utf-8'),
              ensure_ascii=False, separators=(',', ':'))

from collections import Counter
print(Counter(e['type'] for e in entries))
print(len(entries), 'entries ·',
      sum(len(e['senses']) for e in entries), 'senses ·',
      sum(len(s['eg']) for e in entries for s in e['senses']), 'examples ·',
      len(readings), 'passages')
print('dataset.json', round(os.path.getsize(os.path.join(HERE, 'dataset.json')) / 1024), 'KB')
